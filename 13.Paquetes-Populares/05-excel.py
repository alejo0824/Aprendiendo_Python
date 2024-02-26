import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

hoja = wb.active

# wb.create_sheet("Hoja 3")

# hoja3 = wb["Hoja 3"]
# hoja3.title = "Nuevo Titulo"

print(
    hoja.max_row,
    hoja.max_column,
)

# Accediendo a las celdas
celda = hoja["A1"]
celda.value = "Nombre Completo"
print(celda.value)

celda2 = hoja.cell(row=2, column=1)
# print(
#     celda2.value,
#     celda2.row,
#     celda2.column,
#     celda2.coordinate
# )

# Forma de imprimir los valores y posiciones de un excel
for fila in range(1, hoja.max_row+1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        # print(fila, columna, celda.value)

columna = hoja["A"]  # Acceder a  toda una columna
fila = hoja["1"]  # Acceder a una toda Fila

hoja.append([1, 2, 3])  # Agregar registros

hoja.delete_rows(1, 1)  # Eliminar Fila, desde donde hasta dodne

wb.save("Nuevo_Excel.xlsx")
